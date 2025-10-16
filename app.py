from flask import Flask, render_template, request, redirect, url_for, session, flash,  send_file
import sqlite3
import pandas as pd
from datetime import datetime, timedelta
import os
import io
from io import BytesIO

app = Flask(__name__)
app.secret_key = "segredo123"
app.config["MARCACAO_ABERTA"] = False


# Banco de dados SQLite
def init_db():
    with sqlite3.connect("database.db") as conn:
        c = conn.cursor()
        c.execute("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email TEXT UNIQUE NOT NULL,
                qra TEXT NOT NULL,
                posto_grad TEXT NOT NULL,
                re TEXT NOT NULL,
                curso TEXT NOT NULL,
                pelotao INTEGER NOT NULL,
                senha TEXT NOT NULL,
                is_admin INTEGER DEFAULT 0
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS refeicoes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                data TEXT NOT NULL,
                cafe INTEGER DEFAULT 0,
                almoco INTEGER DEFAULT 0,
                janta INTEGER DEFAULT 0,
                FOREIGN KEY(user_id) REFERENCES usuarios(id)
            )
        """)
        conn.commit()


init_db()

# Página inicial de login
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form["email"]
        senha = request.form["senha"]
        if email == "admin" and senha == "admin":
            session["admin"] = True
            return redirect(url_for("admin"))

        with sqlite3.connect("database.db") as con:
            cur = con.cursor()
            cur.execute("SELECT * FROM usuarios WHERE email=? AND senha=?", (email, senha))
            user = cur.fetchone()
            if user:
                session["user_id"] = user[0]
                session["qra"] = user[2]
                return redirect(url_for("dashboard"))
            else:
                return render_template("login.html", erro="Credenciais inválidas.")
    return render_template("login.html")

# Criar conta
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        email = request.form["email"]
        qra = request.form["qra"]
        posto_grad = request.form["posto_grad"]
        re = request.form["re"]
        curso = request.form["curso"]
        pelotao = request.form["pelotao"]
        senha = request.form["senha"]

        try:
            with sqlite3.connect("database.db") as con:
                c = con.cursor()
                c.execute("""
                    INSERT INTO usuarios (email, qra, posto_grad, re, curso, pelotao, senha)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (email, qra, posto_grad, re, curso, pelotao, senha))
                con.commit()
            flash("Conta criada com sucesso! Faça login.", "success")
            return redirect(url_for("login"))
        except sqlite3.IntegrityError:
            erro = "Este email já está cadastrado."
            return render_template("register.html", erro=erro)

    return render_template("register.html")


from datetime import datetime, timedelta

@app.route("/exportar", methods=["GET", "POST"])
def exportar():
    if "admin" not in session:
        return redirect(url_for("login"))

    dados = None
    pelotao_filtro = None

    with sqlite3.connect("database.db") as con:
        cur = con.cursor()
        cur.execute("SELECT DISTINCT pelotao FROM usuarios")
        pelotoes = [row[0] for row in cur.fetchall()]

        if request.method == "POST":
            pelotao_filtro = request.form.get("pelotao")

            query = """
                SELECT u.qra, u.re, u.pelotao, r.data, r.cafe, r.almoco, r.janta
                FROM refeicoes r
                JOIN usuarios u ON r.user_id = u.id
                WHERE (? IS NULL OR u.pelotao = ?)
            """
            cur.execute(query, (pelotao_filtro, pelotao_filtro))
            dados = cur.fetchall()

            if not dados:
                flash("Nenhum dado encontrado para exportar.", "warning")
                return render_template("exportar.html", pelotoes=pelotoes)

            # === MONTA DATAFRAME ===
            df = pd.DataFrame(dados, columns=["re", "qra", "pelotao", "data", "cafe", "almoco", "janta"])

            # === FILTRA APENAS OS PRÓXIMOS 7 DIAS ===
            df["data"] = pd.to_datetime(df["data"]).dt.date
            hoje = datetime.now().date()
            limite = hoje + timedelta(days=7)
            df = df[(df["data"] >= hoje) & (df["data"] <= limite)]

            if df.empty:
                flash("Nenhum registro encontrado nos próximos 7 dias.", "warning")
                return render_template("exportar.html", pelotoes=pelotoes)

            # === FORMATAÇÃO DAS DATAS ===
            df["data"] = df["data"].apply(lambda d: d.strftime("%d/%m/%Y"))
            dias = sorted(df["data"].unique(), key=lambda x: datetime.strptime(x, "%d/%m/%Y"))

            # === CRIA PLANILHA ESTILIZADA ===
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = f"Pelotão {pelotao_filtro}" if pelotao_filtro else "Todos os Pelotões"

            # === ESTILOS ===
            bold_center = Font(bold=True, color="FFFFFF")
            bold_center_black = Font(bold=True, color="000000")
            center = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style="thin", color="555555"),
                right=Side(style="thin", color="555555"),
                top=Side(style="thin", color="555555"),
                bottom=Side(style="thin", color="555555")
            )

            # Cabeçalho principal - Amarelo dourado
            fill_header = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

            # Subcabeçalhos - Cinza médio
            fill_subheader = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

            # Linhas de dados - Cinza claro
            fill_data = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

            # === CABEÇALHOS FIXOS ===
            ws["A1"], ws["B1"] = "RE", "QRA"
            for cell in ["A1", "B1"]:
                ws[cell].font = bold_center
                ws[cell].alignment = center
                ws[cell].fill = fill_header
                ws[cell].border = border

            # === CABEÇALHOS DE DATAS ===
            col = 3  # começa na coluna C
            for dia in dias:
                ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 5)
                ws.cell(row=1, column=col, value=dia)
                ws.cell(row=1, column=col).font = bold_center
                ws.cell(row=1, column=col).alignment = center
                ws.cell(row=1, column=col).fill = fill_header
                ws.cell(row=1, column=col).border = border

                # Subcabeçalhos
                ws.cell(row=2, column=col, value="C")
                ws.cell(row=2, column=col + 2, value="A")
                ws.cell(row=2, column=col + 4, value="J")

                for offset in [0, 2, 4]:
                    ws.cell(row=2, column=col + offset).font = bold_center_black
                    ws.cell(row=2, column=col + offset).alignment = center
                    ws.cell(row=2, column=col + offset).fill = fill_subheader
                    ws.cell(row=2, column=col + offset).border = border

                for offset in [1, 3, 5]:
                    ws.cell(row=2, column=col + offset).fill = fill_subheader
                    ws.cell(row=2, column=col + offset).border = border

                col += 6

            # === PREENCHER DADOS ===
            usuarios = df.drop_duplicates(subset=["re", "qra"])
            start_row = 3

            for i, (_, usuario) in enumerate(usuarios.iterrows(), start=start_row):
                re = usuario["re"]
                qra = usuario["qra"]

                ws.cell(row=i, column=1, value=re)
                ws.cell(row=i, column=2, value=qra)

                for j, dia in enumerate(dias):
                    base_col = 3 + j * 6
                    linha = df[(df["re"] == re) & (df["data"] == dia)]
                    if not linha.empty:
                        dados = linha.iloc[0]
                        if dados["cafe"] == 1:
                            ws.cell(row=i, column=base_col, value=1)
                        if dados["almoco"] == 1:
                            ws.cell(row=i, column=base_col + 2, value=1)
                        if dados["janta"] == 1:
                            ws.cell(row=i, column=base_col + 4, value=1)

                    for c in range(base_col, base_col + 6):
                        ws.cell(row=i, column=c).alignment = center
                        ws.cell(row=i, column=c).border = border
                        ws.cell(row=i, column=c).fill = fill_data

            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 15
            for c in range(3, col):
                ws.column_dimensions[ws.cell(row=2, column=c).column_letter].width = 5

            nome_arquivo = f"relatorio_refeicoes_pelotao_{pelotao_filtro or 'todos'}.xlsx"
            caminho = os.path.join("exports", nome_arquivo)
            os.makedirs("exports", exist_ok=True)
            wb.save(caminho)

            return send_file(caminho, as_attachment=True)

    return render_template("exportar.html", pelotoes=pelotoes)
@app.route("/exportar_total", methods=["POST"])
def exportar_total():
    if "admin" not in session:
        return redirect(url_for("login"))

    with sqlite3.connect("database.db") as con:
        cur = con.cursor()
        cur.execute("""
            SELECT data, 
                   SUM(cafe) as total_cafe, 
                   SUM(almoco) as total_almoco, 
                   SUM(janta) as total_janta
            FROM refeicoes
            GROUP BY data
            ORDER BY data
        """)
        dados = cur.fetchall()

    if not dados:
        return "<script>alert('Nenhum dado encontrado para exportar.'); window.location.href='/admin';</script>"

    # === MONTA DATAFRAME ===
    df = pd.DataFrame(dados, columns=["Data", "Café da Manhã", "Almoço", "Janta"])
    df["Data"] = pd.to_datetime(df["Data"]).dt.strftime("%d/%m/%Y")

    # === CRIA PLANILHA ESTILIZADA ===
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo de Refeições"

    # === ESTILOS ===
    bold_center = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin", color="555555"),
        right=Side(style="thin", color="555555"),
        top=Side(style="thin", color="555555"),
        bottom=Side(style="thin", color="555555")
    )

    fill_header = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    fill_data = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # === CABEÇALHOS ===
    for col, nome in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col, value=nome)
        cell.font = bold_center
        cell.alignment = center
        cell.fill = fill_header
        cell.border = border

    # === PREENCHER DADOS ===
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, valor in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=valor)
            c.alignment = center
            c.border = border
            c.fill = fill_data

    # Ajusta largura das colunas
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    nome_arquivo = "resumo_refeicoes_geral.xlsx"
    caminho = os.path.join("exports", nome_arquivo)
    os.makedirs("exports", exist_ok=True)
    wb.save(caminho)

    return send_file(caminho, as_attachment=True)

# Painel do admin
@app.route("/admin")
def admin():
    if "admin" not in session:
        return redirect(url_for("login"))

    # pega e remove a mensagem da sessão (se existir)
    status_msg = session.pop("status_msg", None)
    with sqlite3.connect("database.db") as con:
        df = pd.read_sql_query("""
            SELECT u.qra, u.re, r.data, r.cafe, r.almoco, r.janta
            FROM refeicoes r
            JOIN usuarios u ON r.user_id = u.id
        """, con)

    if df.empty:
        return render_template("admin.html", tabela_html="<p class='text-center mt-5'>Nenhum dado encontrado.</p>")

    # Converter datas para formato DD/MM/AAAA
    df["data"] = pd.to_datetime(df["data"]).dt.strftime("%d/%m/%Y")

    # Dias únicos ordenados
    dias = sorted(df["data"].unique(), key=lambda x: datetime.strptime(x, "%d/%m/%Y"))

    # Criar lista de usuários únicos
    usuarios = df[["re", "qra"]].drop_duplicates().reset_index(drop=True)

    # Criar estrutura da tabela
    tabela = pd.DataFrame()
    tabela["RE"] = usuarios["re"]
    tabela["QRA"] = usuarios["qra"]

    # Adicionar colunas C, A, J e colunas em branco entre elas para cada dia
    for dia in dias:
        tabela[f"{dia} - C"] = ""
        tabela[f"{dia} - "] = ""  # branco
        tabela[f"{dia} - A"] = ""
        tabela[f"{dia} -  "] = ""  # branco
        tabela[f"{dia} - J"] = ""
        tabela[f"{dia} -   "] = ""  # branco extra

    # Preencher as refeições
    for _, row in df.iterrows():
        re = row["re"]
        dia = row["data"]

        # localizar a linha correspondente ao usuário
        i = tabela.index[tabela["RE"] == re]
        if not i.empty:
            i = i[0]
            if row["cafe"] == 1:
                tabela.loc[i, f"{dia} - C"] = 1
            if row["almoco"] == 1:
                tabela.loc[i, f"{dia} - A"] = 1
            if row["janta"] == 1:
                tabela.loc[i, f"{dia} - J"] = 1

    # Converter DataFrame em tabela HTML formatada
    tabela_html = tabela.to_html(classes="table table-bordered text-center align-middle", index=False, border=0)

    return render_template("admin.html", tabela_html=tabela_html, status_msg=status_msg)

@app.route("/dashboard", methods=["GET", "POST"])
def dashboard():
    if "user_id" not in session:
        return redirect(url_for("login"))

    user_id = session["user_id"]
    hoje = datetime.now()
    dias = [(hoje + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(7)]

    with sqlite3.connect("database.db") as con:
        cur = con.cursor()

        # Atualiza as refeições se for POST
        if request.method == "POST":
         if not app.config.get("MARCACAO_ABERTA", False):
            flash("Sistema fechado — marcação não permitida.", "warning")
         else:
            for data in dias:
                cafe = 1 if request.form.get(f"cafe_{data}") else 0
                almoco = 1 if request.form.get(f"almoco_{data}") else 0
                janta = 1 if request.form.get(f"janta_{data}") else 0

                cur.execute("SELECT * FROM refeicoes WHERE user_id=? AND data=?", (user_id, data))
                if cur.fetchone():
                    cur.execute("UPDATE refeicoes SET cafe=?, almoco=?, janta=? WHERE user_id=? AND data=?",
                                (cafe, almoco, janta, user_id, data))
                else:
                    cur.execute("INSERT INTO refeicoes (user_id, data, cafe, almoco, janta) VALUES (?, ?, ?, ?, ?)",
                                (user_id, data, cafe, almoco, janta))
            con.commit()

        # Depois de salvar (ou carregar), busca novamente os dados atualizados
        refeicoes = {d: (0, 0, 0) for d in dias}
        for data in dias:
            cur.execute("SELECT cafe, almoco, janta FROM refeicoes WHERE user_id=? AND data=?", (user_id, data))
            r = cur.fetchone()
            if r:
                refeicoes[data] = r

    msg = "Refeições salvas com sucesso!" if request.method == "POST" else None
    return render_template("dashboard.html", dias=dias, refeicoes=refeicoes, msg=msg, marcacao_aberta=app.config["MARCACAO_ABERTA"])

# --- substituir abrir_marcacao / fechar_marcacao por estas versões ---
@app.route("/abrir_marcacao", methods=["POST"])
def abrir_marcacao():
    if "admin" not in session:
        # mensagem de erro via session
        session["status_msg"] = {"type": "error", "text": "Acesso negado."}
        return redirect(url_for("login"))
    app.config["MARCACAO_ABERTA"] = True
    # salva mensagem na sessão para o admin ver
    session["status_msg"] = {"type": "success", "text": "✅ Sistema de marcação LIBERADO com sucesso!"}
    return redirect(url_for("admin"))


@app.route("/fechar_marcacao", methods=["POST"])
def fechar_marcacao():
    if "admin" not in session:
        session["status_msg"] = {"type": "error", "text": "Acesso negado."}
        return redirect(url_for("login"))
    app.config["MARCACAO_ABERTA"] = False
    session["status_msg"] = {"type": "warning", "text": "⚠️ Sistema de marcação FECHADO!"}
    return redirect(url_for("admin"))

# Logout
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))  # Pega a porta do ambiente, padrão 5000
    app.run(host="0.0.0.0", port=port, debug=True)




